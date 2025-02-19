from django.shortcuts import render,redirect
from django.http import HttpResponse
from orders.models import orders,order_items
from products.models import variant
from django.utils import timezone
from datetime import datetime, timedelta
from django.contrib.auth import get_user_model
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, TruncYear,TruncDate
from django.db.models import Sum,F,Count
import json
from reportlab.pdfgen import canvas
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import openpyxl
from openpyxl.styles import Font
import io
from django.contrib.auth.decorators import login_required


# Create your views here.
@login_required(login_url='admin_login')
def admin_dash (request):
    if request.user.is_staff:
        
        order = orders.objects.all().exclude(order_status = 'Cancelled')
        order_count = orders.objects.all().count()
        total_sales = int(sum([ord.total_price for ord in order]))
        product_count = variant.objects.all().count()
        

        best_selling_products = (
        order_items.objects.values(
            "variant__product__name")
        .annotate(total_sold=Sum("quantity"))
        .order_by("-total_sold")[:5]
        )
        
        best_selling_categories = (
            order_items.objects
            .values('variant__product__category__name')  
            .annotate(total_sold=Sum('quantity')) 
            .order_by('-total_sold')[:5]  
        )
        best_selling_Brands = (
        order_items.objects.values(
            "variant__product__brand_name")
        .annotate(total_sold=Sum("quantity"))
        .order_by("-total_sold")[:5]
        )



        # Day-wise Sales Data
        sales_data = orders.objects.annotate(date=TruncDate('created_at')) \
                                   .values('date') \
                                   .annotate(total_sales=Sum('total_price')) \
                                   .order_by('date').exclude(order_status = 'Cancelled')

        # Week-wise Sales Data
        week_sales_data = orders.objects.annotate(week=TruncWeek('created_at')) \
                                    .values('week') \
                                    .annotate(total_sales=Sum('total_price')) \
                                    .order_by('week').exclude(order_status = 'Cancelled')

        # Month-wise Sales Data
        month_sales_data = orders.objects.annotate(month=TruncMonth('created_at')) \
                                        .values('month') \
                                        .annotate(total_sales=Sum('total_price')) \
                                        .order_by('month').exclude(order_status = 'Cancelled')

        # year-wise Sales Data
        year_sales_data = orders.objects.annotate(year=TruncYear('created_at')) \
                                    .values('year') \
                                    .annotate(total_sales=Sum('total_price')) \
                                    .order_by('year').exclude(order_status = 'Cancelled')


        day_labels = [sale['date'].strftime('%Y-%m-%d') for sale in sales_data]  # Date labels for x-axis
        day_data = [float(sale['total_sales']) if sale['total_sales'] is not None else 0.0 for sale in sales_data] 

        # Convert data to labels and sales
        week_labels = [sale['week'].strftime('%Y-%m-%d') for sale in week_sales_data]
        week_data = [float(sale['total_sales']) if sale['total_sales'] is not None else 0.0 for sale in week_sales_data]

        month_labels = [sale['month'].strftime('%Y-%m') for sale in month_sales_data]
        month_data = [float(sale['total_sales']) if sale['total_sales'] is not None else 0.0 for sale in month_sales_data]

        year_labels = [sale['year'].strftime('%Y') for sale in year_sales_data]
        year_data = [float(sale['total_sales']) if sale['total_sales'] is not None else 0.0 for sale in year_sales_data]

        context = {'order_count':order_count,
                    'total_sales':total_sales,
                    'product_count':product_count,
                    'day_labels': json.dumps(day_labels),
                    'day_data': json.dumps(day_data),
                    'week_labels': json.dumps(week_labels),
                    'week_data': json.dumps(week_data),
                    'month_labels': json.dumps(month_labels),
                    'month_data': json.dumps(month_data),
                    'year_labels': json.dumps(year_labels),
                    'year_data': json.dumps(year_data),
                    'best_selling_products': best_selling_products,
                    'best_selling_categories': best_selling_categories,
                    'best_selling_Brands': best_selling_Brands,
                    }


        return render(request, 'admin-dashboard.html',context)
    else:
        return redirect('user_login')

@login_required(login_url='admin_login')
def sales_report(request):
    if request.user.is_staff:
        option = request.GET.get('date_range', 'month')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
    
        start_date, end_date = get_date_range(option, start_date, end_date)
        print('start_date:',start_date,"end_date",end_date)
        order_list = orders.objects.filter(created_at__range=[start_date, end_date]).order_by('-created_at')
        Orders = orders.objects.filter(created_at__range=[start_date, end_date]).exclude(order_status = 'Cancelled').order_by('-created_at')
        print(Orders)
        total_orders = Orders.count()
        total_sales_amount = Orders.aggregate(Sum('total_price'))['total_price__sum'] or 0
        discounted_amounts = Orders.aggregate(Sum('discount_price'))['discount_price__sum'] or 0
        total_discount = total_sales_amount - discounted_amounts
        

        

        return render(request, 'sales_report.html', {
            'order_list': order_list,
            'total_orders': total_orders,
            'total_sales_amount': total_sales_amount,
            'total_discount': total_discount,
        })
    else:
        return redirect('user_login')


def get_date_range(option, start_date=None, end_date=None):
    print('this is the option:',option)
    today = datetime.today().date()
    tomorrow = today + timedelta(days=1)
    if option == 'day':
        return today, tomorrow
    elif option == 'week':
        end = today - timedelta(days=today.weekday())  # Start of the week
        start = end - timedelta(days=6)
        print('this is dates :',start,end)
        return start, end
    elif option == 'month':
        start = today.replace(day=1)
        end = (start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        return start, end
    elif option == 'custom':
        try:
            # Convert string inputs to date objects if provided
            start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else today
            end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else today
            print('Custom range:', start, end)
            return start, end
        except ValueError:
            print("Invalid date format. Please provide YYYY-MM-DD format.")
            return today, today
    else:
        return today, today
    
  
def generate_pdf_report(orders, total_orders, total_sales_amount, total_discount, start_date, end_date):
    # Create a buffer to hold the PDF data
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    
    # Title and header section
    c.setFont("Helvetica-Bold", 16)
    c.drawString(200, 750, "Sales Report")
    c.setFont("Helvetica", 12)
    c.drawString(200, 730, f"Seaside Mobile Phones")
    c.drawString(200, 710, f"Email: seasidemobilephones@gmail.com")
    
    # Date range (custom range, if applicable)
    if start_date and end_date:
        date_range_text = f"Date Range: {start_date} to {end_date}"
    else:
        date_range_text = f"Date Range: {start_date} to {end_date}"

    c.drawString(200, 690, date_range_text)
    c.drawString(200, 670, f"Total Orders: {total_orders}")
    c.drawString(200, 650, f"Total Sales: {total_sales_amount}")
    c.drawString(200, 630, f"Total Discounts: {total_discount}")

    # Table header
    c.setFont("Helvetica-Bold", 10)
    c.drawString(50, 600, "Order ID")
    c.drawString(150, 600, "Order Date")
    c.drawString(250, 600, "Status")
    c.drawString(350, 600, "Total Price")
    c.drawString(450, 600, "Discount Price")

    # Table data
    y_position = 580
    c.setFont("Helvetica", 10)
    for order in orders:
        c.drawString(50, y_position, str(order.order_id))
        c.drawString(150, y_position, str(order.created_at.date()))
        c.drawString(250, y_position, order.order_status)
        c.drawString(350, y_position, str(order.total_price))
        c.drawString(450, y_position, str(order.discount_price))
        y_position -= 20  # Move to the next line

    # Finalize and return the response
    c.showPage()
    c.save()
    buffer.seek(0)

    # Return PDF response
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'
    return response


def generate_excel_report(orders, total_orders, total_sales_amount, total_discount, start_date, end_date):
    # Create an Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sales Report"

    # Title and header section
    sheet['A1'] = "Sales Report"
    sheet['A2'] = f"Seaside Mobile Phones"
    sheet['A3'] = f"Email: seasidemobilephones@gmail.com"
    sheet['A4'] = f"Date Range: {start_date} to {end_date}"
    sheet['A5'] = f"Total Orders: {total_orders}"
    sheet['A6'] = f"Total Sales: {total_sales_amount}"
    sheet['A7'] = f"Total Discounts: {total_discount}"

    # Table headers
    headers = ["Order ID", "Order Date", "Status", "Total Price", "Discount Price"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=8, column=col_num, value=header)
        sheet.cell(row=8, column=col_num).font = Font(bold=True)

    # Populate the table with order data
    for row_num, order in enumerate(orders, 9):
        sheet.cell(row=row_num, column=1, value=order.order_id)
        sheet.cell(row=row_num, column=2, value=order.created_at.strftime('%Y-%m-%d'))
        sheet.cell(row=row_num, column=3, value=order.order_status)
        sheet.cell(row=row_num, column=4, value=order.total_price)
        sheet.cell(row=row_num, column=5, value=order.discount_price)
        sheet.cell(row=row_num, column=6, value=order.discount_price)

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
    workbook.save(response)
    return response


@login_required(login_url='admin_login')
def generate_pdf_report_view(request):
    if request.user.is_staff:
        option = request.GET.get('date_range', 'month')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Fetch filtered data
        start_date, end_date = get_date_range(option, start_date, end_date)
        Orders = orders.objects.filter(created_at__range=[start_date, end_date]).exclude(order_status='Cancelled')

        total_orders = Orders.count()
        total_sales_amount = Orders.aggregate(Sum('total_price'))['total_price__sum'] or 0
        discounted_amounts = Orders.aggregate(Sum('discount_price'))['discount_price__sum'] or 0
        total_discount = total_sales_amount - discounted_amounts

        return generate_pdf_report(Orders, total_orders, total_sales_amount, total_discount, start_date, end_date)
    else:
        return HttpResponse("Unauthorized", status=401)
    
    
@login_required(login_url='admin_login')
def generate_excel_report_view(request):
    if request.user.is_staff:
        option = request.GET.get('date_range', 'month')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Fetch filtered data
        start_date, end_date = get_date_range(option, start_date, end_date)
        Orders = orders.objects.filter(created_at__range=[start_date, end_date]).exclude(order_status='Cancelled')

        total_orders = Orders.count()
        total_sales_amount = Orders.aggregate(Sum('total_price'))['total_price__sum'] or 0
        discounted_amounts = Orders.aggregate(Sum('discount_price'))['discount_price__sum'] or 0
        total_discount = total_sales_amount - discounted_amounts

        return generate_excel_report(Orders, total_orders, total_sales_amount, total_discount, start_date, end_date)
    else:
        return HttpResponse("Unauthorized", status=401)
    

